import pytesseract
pytesseract.pytesseract.tesseract_cmd = r"./flask-server/module1/Tesseract-OCR/tesseract.exe"
import re
import openpyxl
from openpyxl.styles import PatternFill , Alignment
from module1.predict import get_predict

############################################EXCEL AND PREDICT FUNCTIONS#####################
def get_shapes_labels(shape_cells , alreadyMadeOCR = False): ###TAKE image cells ang get name from prediction models and use one of two models (my model ,Teseract)
    cnt=0
    total_labels=[]
    for i in range(len(shape_cells)):
        label=[]
        for j in range(3):
            if cnt % 3 == 0:
                if alreadyMadeOCR :
                    label.append(pytesseract.image_to_string(shape_cells[i][j]))          
                else:
                    label.append(get_predict(shape_cells[i][j],"DIGITS"))        
            else:
                label.append(get_predict(shape_cells[i][j],"SHAPES"))
            cnt=cnt+1
            
        total_labels.append(label)

    return total_labels

def get_excel_labels(total_labels): ## get labels and update total labels to be ready for excel 


    for i in range(len(total_labels)):
       
        for j in range(3):
            if total_labels[i][j] == "Horizontal1":
                total_labels[i][j] = 4

            elif total_labels[i][j] == "Horizontal2":
                total_labels[i][j] = 3

            elif total_labels[i][j] == "Horizontal3":
                total_labels[i][j] = 2

            elif total_labels[i][j] == "Horizontal4":
                total_labels[i][j] = 1

            elif total_labels[i][j] == "Vertical1":
                total_labels[i][j] = 1

            elif total_labels[i][j] == "Vertical2":
                total_labels[i][j] = 2

            elif total_labels[i][j] == "Vertical3":
                total_labels[i][j] = 3

            elif total_labels[i][j] == "Vertical4":
                total_labels[i][j] = 4

            elif total_labels[i][j] == "Vertical5":
                total_labels[i][j] = 5

            elif total_labels[i][j] == "Check":
                total_labels[i][j] = 5

            elif total_labels[i][j] == "Square":
                total_labels[i][j] = 0

            elif total_labels[i][j] == "Question":
                total_labels[i][j] = "?"

            elif total_labels[i][j] == "Empty":
                total_labels[i][j] = " "

    return total_labels
            
def get_english_names(english_name_cells): ## get english names by teseract for excel

    english_names = []
    for cell in english_name_cells: 
        english_names.append(re.sub('[^a-zA-Z ]', '', pytesseract.image_to_string(cell)))
    return english_names

def get_ids(id_cells): ## get ids by teseract for excel
    id_numbers = []
    for cell in id_cells:
        id_numbers.append(re.sub(r'[^0-9]', '', pytesseract.image_to_string(cell)))
    return id_numbers
 
def write_to_excel(file_path, id_numbers, english_names,total_labels): ##
    # Create a new workbook or load an existing one
    workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    sheet.append(["Code","Name",1,2,3])
    red = PatternFill(patternType="solid",fgColor="FF0000")
    
    # Write values to the first column
    cnt = 0
    for id in id_numbers:

        sheet.append([id,english_names[cnt], total_labels[cnt][0], total_labels[cnt][1] , total_labels[cnt][2] ])
        if total_labels[cnt][0] == "?":
            sheet["C"+str(cnt+2)].fill = red
            sheet["C"+str(cnt+2)].value = " "
        if total_labels[cnt][1] == "?":
            sheet["D"+str(cnt+2)].fill = red
            sheet["D"+str(cnt+2)].value = " "
        if total_labels[cnt][2] == "?":
            sheet["E"+str(cnt+2)].fill = red
            sheet["E"+str(cnt+2)].value = " "
        cnt = cnt+1
    
    # align and Save the workbook
    sheet = align_sheet(sheet,file_path)
    workbook.save(file_path)
    
def align_sheet(sheet,file_path):

    for row in sheet.iter_rows():
        for cell in row:
            # Apply center alignment to each cell
            cell.alignment = Alignment(horizontal='center')

    for letter in ['A','B']: 
        max_width = 0
        for row_number in range(1,sheet.max_row + 1):
            if len(str(sheet[f'{letter}{row_number}'].value)) > max_width:
                max_width = len(str(sheet[f'{letter}{row_number}'].value))
            sheet.column_dimensions[letter].width = max_width + 1
        sheet.column_dimensions[letter].width = max_width + 1

   
